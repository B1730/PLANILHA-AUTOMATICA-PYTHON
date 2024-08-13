import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
drive = webdriver.Chrome()
    # utilizei por ser gratuito
drive.get('https://consultcpf-devaprender.netlify.app/')
# Para entrar no excel:
planilia_dos_clientes = openpyxl.load_workbook('dados_clientes1.xlsx')
paginas_dos_clientes = planilia_dos_clientes['Dados_dos_clientes_pg1']
# Pega as informações
for linha in paginas_dos_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    sleep(3)
    pesquisa = drive.find_element(By.XPATH, "//input[@id='cpfInput']")#O id do cpf (digite o cpf)

    sleep(1) #Coloco tempo so pra não travar
    pesquisa.clear()
    pesquisa.send_keys(cpf)
    sleep(1) 
    #Ver se esta atrasado
    botao_pesquisa = pesquisa = drive.find_element(By.XPATH, '//button[@class="btn btn-custom btn-lg btn-block mt-3"]')
    sleep(1)
    botao_pesquisa.click()
    sleep(3)
    #Ver se está em dia ou não
    analise = drive.find_element(By.XPATH, "//span[@id='statusLabel']")
    if analise.text == 'em dia':
        data_pagament = drive.find_element(By.XPATH, "//p[@id='paymentDate']")
        metodo_pagamentp = drive.find_element(By.XPATH, "//p[@id='paymentMethod']")
        data_pagamento_feito= data_pagament.text.split()[3]
        metodo_pagamento_feito = metodo_pagamentp.text.split()[3]

        planilia_dos_fechamentos = openpyxl.load_workbook('planilha fechamento.xlsx')
        fechamento = planilia_dos_fechamentos['Sheet1']

        fechamento.append([nome,valor,cpf,vencimento,'em dia',data_pagamento_feito,metodo_pagamento_feito])

        planilia_dos_fechamentos.save('planilha fechamento.xlsx')#pra salvar a planilia
    else:
        planilia_dos_fechamentos = openpyxl.load_workbook('planilha fechamento.xlsx')
        fechamento = planilia_dos_fechamentos['Sheet1']

        fechamento.append([nome,valor,cpf,vencimento,'pendente'])
        planilia_dos_fechamentos.save('planilha fechamento.xlsx')
        